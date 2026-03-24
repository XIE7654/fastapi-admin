"""
短信日志控制器
"""
import io
from typing import List, Optional
from datetime import datetime
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from app.core.database import get_db
from app.core.dependencies import check_permission
from app.module.system.model.user import User
from app.module.system.service.sms_log import SmsLogService
from app.common.response import success, page_success

router = APIRouter()


# 短信类型字典
SMS_TEMPLATE_TYPE_DICT = {1: "验证码", 2: "通知", 3: "营销"}

# 用户类型字典
USER_TYPE_DICT = {1: "管理员", 2: "会员"}

# 发送状态字典
SMS_SEND_STATUS_DICT = {0: "发送中", 10: "发送成功", 20: "发送失败", 30: "不发送"}

# 接收状态字典
SMS_RECEIVE_STATUS_DICT = {0: "等待接收", 10: "接收成功", 20: "接收失败"}


@router.get("/page", summary="获得短信日志分页")
async def get_sms_log_page(
    page_no: int = Query(1, ge=1, alias="pageNo", description="页码"),
    page_size: int = Query(10, ge=1, le=100, alias="pageSize", description="每页数量"),
    channel_id: int = Query(None, alias="channelId", description="渠道编号"),
    template_id: int = Query(None, alias="templateId", description="模板编号"),
    mobile: str = Query(None, description="手机号"),
    send_status: int = Query(None, alias="sendStatus", description="发送状态"),
    receive_status: int = Query(None, alias="receiveStatus", description="接收状态"),
    send_time: List[datetime] = Query(None, alias="sendTime", description="发送时间"),
    receive_time: List[datetime] = Query(None, alias="receiveTime", description="接收时间"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(check_permission("system:sms-log:query")),
):
    """分页查询短信日志"""
    logs, total = await SmsLogService.get_page(
        db,
        page_no=page_no,
        page_size=page_size,
        channel_id=channel_id,
        template_id=template_id,
        mobile=mobile,
        send_status=send_status,
        receive_status=receive_status,
        send_time=send_time,
        receive_time=receive_time,
    )
    return page_success(
        list_data=[
            {
                "id": log.id,
                "channelId": log.channel_id,
                "channelCode": log.channel_code,
                "templateId": log.template_id,
                "templateCode": log.template_code,
                "templateType": log.template_type,
                "templateContent": log.template_content,
                "templateParams": log.template_params,
                "apiTemplateId": log.api_template_id,
                "mobile": log.mobile,
                "userId": log.user_id,
                "userType": log.user_type,
                "sendStatus": log.send_status,
                "sendTime": log.send_time,
                "apiSendCode": log.api_send_code,
                "apiSendMsg": log.api_send_msg,
                "apiRequestId": log.api_request_id,
                "apiSerialNo": log.api_serial_no,
                "receiveStatus": log.receive_status,
                "receiveTime": log.receive_time,
                "apiReceiveCode": log.api_receive_code,
                "apiReceiveMsg": log.api_receive_msg,
                "createTime": log.create_time,
            }
            for log in logs
        ],
        total=total,
        page_no=page_no,
        page_size=page_size,
    )


@router.get("/get", summary="获得短信日志")
async def get_sms_log(
    id: int = Query(..., description="日志编号"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(check_permission("system:sms-log:query")),
):
    """根据ID获取短信日志详情"""
    log = await SmsLogService.get_by_id(db, id)
    if not log:
        return success(data=None)
    return success(data={
        "id": log.id,
        "channelId": log.channel_id,
        "channelCode": log.channel_code,
        "templateId": log.template_id,
        "templateCode": log.template_code,
        "templateType": log.template_type,
        "templateContent": log.template_content,
        "templateParams": log.template_params,
        "apiTemplateId": log.api_template_id,
        "mobile": log.mobile,
        "userId": log.user_id,
        "userType": log.user_type,
        "sendStatus": log.send_status,
        "sendTime": log.send_time,
        "apiSendCode": log.api_send_code,
        "apiSendMsg": log.api_send_msg,
        "apiRequestId": log.api_request_id,
        "apiSerialNo": log.api_serial_no,
        "receiveStatus": log.receive_status,
        "receiveTime": log.receive_time,
        "apiReceiveCode": log.api_receive_code,
        "apiReceiveMsg": log.api_receive_msg,
        "createTime": log.create_time,
    })


@router.get("/export-excel", summary="导出短信日志 Excel")
async def export_sms_log_excel(
    channel_id: int = Query(None, alias="channelId", description="渠道编号"),
    template_id: int = Query(None, alias="templateId", description="模板编号"),
    mobile: str = Query(None, description="手机号"),
    send_status: int = Query(None, alias="sendStatus", description="发送状态"),
    receive_status: int = Query(None, alias="receiveStatus", description="接收状态"),
    send_time: List[datetime] = Query(None, alias="sendTime", description="发送时间"),
    receive_time: List[datetime] = Query(None, alias="receiveTime", description="接收时间"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(check_permission("system:sms-log:export")),
):
    """导出短信日志 Excel"""
    # 获取数据
    logs = await SmsLogService.get_list(
        db,
        channel_id=channel_id,
        template_id=template_id,
        mobile=mobile,
        send_status=send_status,
        receive_status=receive_status,
        send_time=send_time,
        receive_time=receive_time,
    )

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "短信日志"

    # 定义表头
    headers = [
        "编号", "短信渠道编号", "短信渠道编码", "模板编号", "模板编码",
        "短信类型", "短信内容", "短信参数", "短信API的模板编号",
        "手机号", "用户编号", "用户类型", "发送状态", "发送时间",
        "短信API发送结果的编码", "短信API发送失败的提示", "短信API发送返回的唯一请求ID",
        "短信API发送返回的序号", "接收状态", "接收时间",
        "API接收结果的编码", "API接收结果的说明", "创建时间"
    ]

    # 设置表头样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row, log in enumerate(logs, 2):
        data_row = [
            log.id,
            log.channel_id,
            log.channel_code,
            log.template_id,
            log.template_code,
            SMS_TEMPLATE_TYPE_DICT.get(log.template_type, log.template_type),
            log.template_content,
            log.template_params,
            log.api_template_id,
            log.mobile,
            log.user_id,
            USER_TYPE_DICT.get(log.user_type, log.user_type) if log.user_type else None,
            SMS_SEND_STATUS_DICT.get(log.send_status, log.send_status),
            log.send_time.strftime("%Y-%m-%d %H:%M:%S") if log.send_time else None,
            log.api_send_code,
            log.api_send_msg,
            log.api_request_id,
            log.api_serial_no,
            SMS_RECEIVE_STATUS_DICT.get(log.receive_status, log.receive_status),
            log.receive_time.strftime("%Y-%m-%d %H:%M:%S") if log.receive_time else None,
            log.api_receive_code,
            log.api_receive_msg,
            log.create_time.strftime("%Y-%m-%d %H:%M:%S") if log.create_time else None,
        ]
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 返回文件流
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=sms_log.xlsx"
        }
    )