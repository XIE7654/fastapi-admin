"""
Excel 导出工具类
"""
import io
from typing import List, Any, Optional, Callable
from datetime import datetime
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from fastapi.responses import StreamingResponse


def _encode_filename(filename: str) -> str:
    """
    编码文件名，支持中文等非 ASCII 字符
    使用 RFC 5987 标准编码 Content-Disposition 头中的文件名
    """
    try:
        # 尝试用 latin-1 编码，如果成功直接使用
        filename.encode('latin-1')
        return f'attachment; filename="{filename}"'
    except UnicodeEncodeError:
        # 包含非 ASCII 字符，使用 URL 编码
        encoded = quote(filename)
        return f"attachment; filename*=UTF-8''{encoded}"


class ExcelUtils:
    """Excel 导出工具类"""

    # 默认样式
    HEADER_FONT = Font(bold=True)
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    @staticmethod
    def export_excel(
        data: List[Any],
        headers: List[str],
        fields: List[str],
        filename: str = "export.xlsx",
        sheet_name: str = "Sheet1",
        converters: Optional[dict] = None,
    ) -> StreamingResponse:
        """
        导出 Excel 文件

        Args:
            data: 数据列表（可以是字典列表或对象列表）
            headers: 表头列表
            fields: 字段名列表（与表头一一对应）
            filename: 文件名
            sheet_name: 工作表名称
            converters: 字段转换器字典，key 为字段名，value 为转换函数

        Returns:
            StreamingResponse: 流式响应
        """
        converters = converters or {}

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = ExcelUtils.HEADER_FONT
            cell.alignment = ExcelUtils.HEADER_ALIGNMENT
            cell.border = ExcelUtils.THIN_BORDER

        # 写入数据
        for row_idx, item in enumerate(data, 2):
            for col, field in enumerate(fields, 1):
                # 获取字段值
                if isinstance(item, dict):
                    value = item.get(field)
                else:
                    value = getattr(item, field, None)

                # 应用转换器
                if field in converters:
                    value = converters[field](value)

                # 处理时间类型
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")

                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = ExcelUtils.THIN_BORDER

        # 自动调整列宽
        ExcelUtils._auto_adjust_column_width(ws)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 返回文件流
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": _encode_filename(filename)
            }
        )

    @staticmethod
    def export_excel_with_row_data(
        data: List[List[Any]],
        headers: List[str],
        filename: str = "export.xlsx",
        sheet_name: str = "Sheet1",
    ) -> StreamingResponse:
        """
        导出 Excel 文件（使用行数据）

        Args:
            data: 行数据列表，每个元素是一行数据的列表
            headers: 表头列表
            filename: 文件名
            sheet_name: 工作表名称

        Returns:
            StreamingResponse: 流式响应
        """
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = ExcelUtils.HEADER_FONT
            cell.alignment = ExcelUtils.HEADER_ALIGNMENT
            cell.border = ExcelUtils.THIN_BORDER

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                # 处理时间类型
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")

                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = ExcelUtils.THIN_BORDER

        # 自动调整列宽
        ExcelUtils._auto_adjust_column_width(ws)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 返回文件流
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": _encode_filename(filename)
            }
        )

    @staticmethod
    def _auto_adjust_column_width(ws):
        """自动调整列宽"""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # 最大宽度限制为 50
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = max(adjusted_width, 10)